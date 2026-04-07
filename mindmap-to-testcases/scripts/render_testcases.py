#!/usr/bin/env python3
"""Render structured test cases JSON into CSV or XLSX.

Usage:
  python scripts/render_testcases.py input.json output.csv
  python scripts/render_testcases.py input.json output.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_COLUMNS = [
    "所属模块",
    "用例标题",
    "前置条件",
    "步骤",
    "预期",
    "优先级",
    "用例类型",
    "适用阶段",
]

ALIASES = {
    "所属模块": ["所属模块", "module"],
    "用例标题": ["用例标题", "title"],
    "前置条件": ["前置条件", "preconditions"],
    "步骤": ["步骤", "steps"],
    "预期": ["预期", "expected"],
    "优先级": ["优先级", "priority"],
    "用例类型": ["用例类型", "case_type"],
    "适用阶段": ["适用阶段", "stage"],
}

COLUMN_WIDTHS = {
    "所属模块": 24,
    "用例标题": 42,
    "前置条件": 36,
    "步骤": 44,
    "预期": 36,
    "优先级": 10,
    "用例类型": 14,
    "适用阶段": 16,
}

CASE_TYPE_MAP = {
    "功能测试": "功能测试",
    "异常测试": "异常测试",
    "边界测试": "边界测试",
    "场景测试": "场景测试",
    "functional": "功能测试",
    "abnormal": "异常测试",
    "boundary": "边界测试",
    "scenario": "场景测试",
}
ENGLISH_CASE_TYPES = {"functional", "abnormal", "boundary", "scenario"}
PLACEHOLDER_RE = re.compile(r"\?{2,}|�")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Render JSON test cases to CSV or XLSX")
    parser.add_argument("input", help="Path to JSON file")
    parser.add_argument("output", help="Path to output .csv or .xlsx file")
    parser.add_argument("--sheet-name", default="testcases", help="Worksheet name when exporting XLSX")
    return parser.parse_args()


def load_cases(path: Path) -> List[Dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, list):
        cases = data
    elif isinstance(data, dict) and isinstance(data.get("cases"), list):
        cases = data["cases"]
    else:
        raise ValueError("input JSON must be a list or an object with a 'cases' list")
    if not all(isinstance(item, dict) for item in cases):
        raise ValueError("each case must be a JSON object")
    return cases


def get_value(case: Dict[str, Any], field: str) -> Any:
    for key in ALIASES[field]:
        if key in case:
            return case[key]
    return ""


def ensure_numbered_lines(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if not isinstance(value, Iterable):
        return str(value)

    lines: List[str] = []
    for index, item in enumerate(value, start=1):
        text = str(item).strip()
        if not text:
            continue
        if re.match(r"^\d+\.\s*", text):
            lines.append(text)
        else:
            lines.append(f"{index}. {text}")
    return "\n".join(lines)


def normalize_priority(value: Any) -> str:
    text = str(value).strip()
    mapping = {
        "P0": "1",
        "P1": "2",
        "P2": "3",
        "p0": "1",
        "p1": "2",
        "p2": "3",
    }
    return mapping.get(text, text)


def normalize_case_type(value: Any) -> str:
    text = str(value).strip()
    return CASE_TYPE_MAP.get(text, text)


def normalize_case(case: Dict[str, Any]) -> Dict[str, str]:
    row: Dict[str, str] = {}
    for field in BASE_COLUMNS:
        value = get_value(case, field)
        if field in {"前置条件", "步骤", "预期"}:
            row[field] = ensure_numbered_lines(value)
        elif field == "优先级":
            row[field] = normalize_priority(value)
        elif field == "用例类型":
            row[field] = normalize_case_type(value)
        else:
            row[field] = str(value).strip()
    return row


def write_csv(rows: List[Dict[str, str]], output: Path) -> None:
    with output.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=BASE_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in BASE_COLUMNS})


def write_xlsx(rows: List[Dict[str, str]], output: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "testcases")[:31]
    ws.append(BASE_COLUMNS)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in rows:
        ws.append([row.get(column, "") for column in BASE_COLUMNS])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_alignment

    for index, column in enumerate(BASE_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output)


def fallback_output_path(output: Path) -> Path:
    candidate = output.with_name(f"{output.stem}.fixed{output.suffix}")
    if not candidate.exists():
        return candidate
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return output.with_name(f"{output.stem}.{timestamp}{output.suffix}")


def contains_suspicious_placeholder(text: str) -> bool:
    if not text:
        return False
    if PLACEHOLDER_RE.search(text):
        return True
    question_count = text.count("?")
    return question_count >= 3 or (question_count >= 2 and question_count / max(len(text), 1) >= 0.2)


def verify_csv(output: Path, expected_count: int) -> None:
    raw = output.read_bytes()
    if not raw.startswith(b"\xef\xbb\xbf"):
        raise ValueError(f"csv export is missing UTF-8 BOM: {output}")

    with output.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    if not rows or rows[0] != BASE_COLUMNS:
        raise ValueError(f"csv header mismatch after export: {output}")
    if len(rows) - 1 != expected_count:
        raise ValueError(f"csv row count mismatch after export: expected {expected_count}, got {len(rows) - 1}")

    case_type_index = BASE_COLUMNS.index("用例类型")
    for row_index, row in enumerate(rows[1:], start=2):
        if len(row) != len(BASE_COLUMNS):
            raise ValueError(f"csv column count mismatch at row {row_index}: {output}")
        if row[case_type_index] in ENGLISH_CASE_TYPES:
            raise ValueError(f"csv contains unnormalized english case type at row {row_index}: {row[case_type_index]}")
        if any(contains_suspicious_placeholder(cell) for cell in row):
            raise ValueError(f"csv contains suspicious placeholder text at row {row_index}")


def verify_xlsx(output: Path, expected_count: int) -> None:
    wb = load_workbook(output, read_only=True)
    try:
        ws = wb.active
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if headers != BASE_COLUMNS:
            raise ValueError(f"xlsx header mismatch after export: {output}")

        row_count = 0
        case_type_index = BASE_COLUMNS.index("用例类型")
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_count += 1
            cells = ["" if cell is None else str(cell) for cell in row]
            if cells[case_type_index] in ENGLISH_CASE_TYPES:
                raise ValueError(f"xlsx contains unnormalized english case type at row {row_index}: {cells[case_type_index]}")
            if any(contains_suspicious_placeholder(cell) for cell in cells):
                raise ValueError(f"xlsx contains suspicious placeholder text at row {row_index}")

        if row_count != expected_count:
            raise ValueError(f"xlsx row count mismatch after export: expected {expected_count}, got {row_count}")
    finally:
        wb.close()


def write_with_fallback(rows: List[Dict[str, str]], output: Path, sheet_name: str) -> Path:
    suffix = output.suffix.lower()
    try:
        if suffix == ".csv":
            write_csv(rows, output)
        elif suffix == ".xlsx":
            write_xlsx(rows, output, sheet_name)
        else:
            raise SystemExit("output extension must be .csv or .xlsx")
        return output
    except PermissionError:
        fallback = fallback_output_path(output)
        if suffix == ".csv":
            write_csv(rows, fallback)
        elif suffix == ".xlsx":
            write_xlsx(rows, fallback, sheet_name)
        return fallback


def verify_output(output: Path, expected_count: int) -> None:
    suffix = output.suffix.lower()
    if suffix == ".csv":
        verify_csv(output, expected_count)
        return
    if suffix == ".xlsx":
        verify_xlsx(output, expected_count)
        return
    raise SystemExit("output extension must be .csv or .xlsx")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)
    rows = [normalize_case(case) for case in load_cases(input_path)]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    actual_output = write_with_fallback(rows, output_path, args.sheet_name)
    verify_output(actual_output, len(rows))

    if actual_output != output_path:
        print(f"Exported {len(rows)} case(s) to {actual_output} (fallback path used because target file was locked)")
    else:
        print(f"Exported {len(rows)} case(s) to {actual_output}")


if __name__ == "__main__":
    main()
