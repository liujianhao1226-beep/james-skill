#!/usr/bin/env python3
import argparse
import csv
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def read_csv(path: Path) -> str:
    lines = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if any(cell.strip() for cell in row if cell):
                lines.append("\t".join(cell.strip() for cell in row))
    return "\n".join(lines)


def read_xlsx(path: Path) -> str:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx files")
    wb = load_workbook(path, read_only=True, data_only=True)
    chunks = []
    for ws in wb.worksheets:
        chunks.append(f"--- SHEET: {ws.title} ---")
        for row in ws.iter_rows(values_only=True):
            values = [str(v).strip() if v is not None else "" for v in row]
            if any(values):
                chunks.append("\t".join(values))
    return "\n".join(chunks)


def read_docx(path: Path) -> str:
    with zipfile.ZipFile(path, "r") as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for para in root.findall(".//w:p", ns):
        texts = [node.text for node in para.findall(".//w:t", ns) if node.text]
        if texts:
            lines.append("".join(texts))
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Flatten requirement sources to UTF-8 text.")
    parser.add_argument("--input", required=True, help="Path to source file")
    parser.add_argument("--output", help="Optional output text path")
    args = parser.parse_args()

    src = Path(args.input)
    suffix = src.suffix.lower()

    if suffix in {".md", ".txt"}:
        text = read_text(src)
    elif suffix == ".csv":
        text = read_csv(src)
    elif suffix == ".xlsx":
        text = read_xlsx(src)
    elif suffix == ".docx":
        text = read_docx(src)
    elif suffix == ".pdf":
        raise RuntimeError("PDF extraction is not bundled. Convert PDF to text/docx first.")
    else:
        raise RuntimeError(f"Unsupported file type: {suffix}")

    if args.output:
        Path(args.output).write_text(text, encoding="utf-8")
    else:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        print(text)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
